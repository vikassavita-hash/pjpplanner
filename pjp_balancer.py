import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime

def haversine_dist(lat1, lon1, lat2, lon2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians 
    lat1_r, lon1_r, lat2_r, lon2_r = map(np.radians, [lat1, lon1, lat2, lon2])

    # haversine formula 
    dlat = lat2_r - lat1_r
    dlon = lon2_r - lon1_r
    a = np.sin(dlat/2.0)**2 + np.cos(lat1_r) * np.cos(lat2_r) * np.sin(dlon/2.0)**2
    c = 2 * np.arcsin(np.sqrt(a))
    r = 6371 # Radius of earth in kilometers
    return c * r

def generate_dynamic_pjp(df_input, num_merch, format_type="excel"):
    """
    Executes the dynamic VRP and PJP balancing pipeline:
    1. Clean and validate inputs
    2. Geographic clustering using KMeans (weighted by frequency)
    3. Greedy Haversine-based workload re-balancing (+/- 5% tolerance)
    4. Rhythmic binary date matrix spacing across 24 working days
    5. Openpyxl formatted multi-tab Excel workbook generation
    """
    # ----------------------------------------------------
    # 1. DYNAMIC SCRIPT ENGINE & INPUTS
    # ----------------------------------------------------
    df = df_input.copy()
    
    # Standardize column mappings (case-insensitive & whitespace trimmed)
    rename_map = {}
    for col in df.columns:
        c_lower = str(col).strip().lower()
        if c_lower in ['outlet code', 'outlet_code', 'code']:
            rename_map[col] = 'Outlet code'
        elif c_lower in ['outlet name', 'outlet_name', 'name']:
            rename_map[col] = 'Outlet name'
        elif c_lower in ['beat']:
            rename_map[col] = 'Beat'
        elif c_lower in ['lat', 'latitude', 'lat/long', 'lat_long']:
            rename_map[col] = 'LAT'
        elif c_lower in ['long', 'longitude', 'lng']:
            rename_map[col] = 'LONG'
        elif c_lower in ['actaul frequency', 'actaul_frequency', 'actual frequency', 'actual_frequency', 'frequency', 'freq', 'act frequency']:
            rename_map[col] = 'Actaul Frequency'
            
    df.rename(columns=rename_map, inplace=True)
    
    required_cols = ['Outlet code', 'Outlet name', 'Beat', 'LAT', 'LONG', 'Actaul Frequency']
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Input file is missing required columns: {missing_cols}")
        
    # Clean missing values and cast types strictly
    df.dropna(subset=['Outlet code', 'LAT', 'LONG', 'Actaul Frequency'], inplace=True)
    
    df['Outlet code'] = df['Outlet code'].astype(int)
    df['Actaul Frequency'] = df['Actaul Frequency'].astype(int)
    df['LAT'] = df['LAT'].astype(np.float64)
    df['LONG'] = df['LONG'].astype(np.float64)
    
    # Exclude any 0-frequency or negative-frequency outlets
    df = df[df['Actaul Frequency'] > 0].reset_index(drop=True)
    
    if len(df) == 0:
        raise ValueError("No valid outlet rows containing positive frequency and coordinates found.")
    
    # ----------------------------------------------------
    # 2. COORDINATE-LEVEL GEOGRAPHIC CLUSTERING
    # ----------------------------------------------------
    # KMeans territory clustering weighted by Actual Frequency
    coords = df[['LAT', 'LONG']].values
    freq_weights = df['Actaul Frequency'].values
    
    # Ensure num_merch is within bounds
    num_merch = max(9, int(num_merch))
    if num_merch > len(df):
        num_merch = len(df)
        
    kmeans = KMeans(n_clusters=num_merch, random_state=42, n_init=10)
    df['merch_id'] = kmeans.fit_predict(coords, sample_weight=freq_weights)
    centroids = kmeans.cluster_centers_
    
    # ----------------------------------------------------
    # GREEDY RE-BALANCING LOOP (+/- 5% target)
    # ----------------------------------------------------
    total_freq = df['Actaul Frequency'].sum()
    global_mean = total_freq / num_merch
    min_target = 0.95 * global_mean
    max_target = 1.05 * global_mean
    
    # Calculate Haversine distance matrix from all outlets to all centroids
    num_outlets = len(df)
    dist_matrix = np.zeros((num_outlets, num_merch))
    for m in range(num_merch):
        dist_matrix[:, m] = haversine_dist(
            df['LAT'].values, df['LONG'].values,
            centroids[m, 0], centroids[m, 1]
        )
        
    assignments = df['merch_id'].values.copy()
    
    def get_merch_loads(assigns):
        loads = np.zeros(num_merch)
        np.add.at(loads, assigns, freq_weights)
        return loads

    # Iterative re-balancing to enforce capacity bounds
    max_iters = 1500
    for _ in range(max_iters):
        loads = get_merch_loads(assignments)
        overloaded = np.where(loads > max_target)[0]
        underloaded = np.where(loads < min_target)[0]
        
        if len(overloaded) == 0 and len(underloaded) == 0:
            break  # Perfectly balanced
            
        moved = False
        
        # 1. Try to shift outlets away from overloaded merchandisers
        if len(overloaded) > 0:
            # Pick the most overloaded
            from_m = overloaded[np.argmax(loads[overloaded])]
            from_outlets = np.where(assignments == from_m)[0]
            
            best_move = None
            min_dist = float('inf')
            
            for o_idx in from_outlets:
                o_freq = freq_weights[o_idx]
                # Try to move this outlet to a merchandiser that won't exceed target
                for to_m in range(num_merch):
                    if to_m == from_m:
                        continue
                    if loads[to_m] + o_freq <= max_target:
                        dist = dist_matrix[o_idx, to_m]
                        if dist < min_dist:
                            min_dist = dist
                            best_move = (o_idx, to_m)
                            
            if best_move:
                o_idx, to_m = best_move
                assignments[o_idx] = to_m
                moved = True
                
        # 2. Try to pull outlets into underloaded merchandisers
        if not moved and len(underloaded) > 0:
            # Pick the most underloaded
            to_m = underloaded[np.argmin(loads[underloaded])]
            
            best_move = None
            min_dist = float('inf')
            
            for from_m in range(num_merch):
                if from_m == to_m:
                    continue
                from_outlets = np.where(assignments == from_m)[0]
                for o_idx in from_outlets:
                    o_freq = freq_weights[o_idx]
                    # Make sure pulling doesn't drop the sender below min target, 
                    # or at least the sender is larger than the receiver
                    if loads[from_m] - o_freq >= min_target or loads[from_m] > loads[to_m] + o_freq:
                        dist = dist_matrix[o_idx, to_m]
                        if dist < min_dist:
                            min_dist = dist
                            best_move = (o_idx, to_m)
                            
            if best_move:
                o_idx, to_m = best_move
                assignments[o_idx] = to_m
                moved = True
                
        # 3. Fallback: Reduce workload variance if stuck
        if not moved:
            from_m = np.argmax(loads)
            to_m = np.argmin(loads)
            from_outlets = np.where(assignments == from_m)[0]
            
            best_move = None
            min_dist = float('inf')
            for o_idx in from_outlets:
                o_freq = freq_weights[o_idx]
                if loads[from_m] - loads[to_m] > 2 * o_freq:
                    dist = dist_matrix[o_idx, to_m]
                    if dist < min_dist:
                        min_dist = dist
                        best_move = (o_idx, to_m)
            if best_move:
                o_idx, to_m = best_move
                assignments[o_idx] = to_m
                moved = True
            else:
                break # No further moves can improve load distribution
                
    df['merch_id'] = assignments
    df['Merchandiser Name'] = df['merch_id'].apply(lambda x: f"Merchandiser {x + 1:02d}")
    
    # ----------------------------------------------------
    # 3. BINARY DATE MATRIX SPACING LOGIC
    # ----------------------------------------------------
    # Generate 24 working days starting from June 2026 (excluding Sundays)
    start_date = datetime.date(2026, 6, 1)
    date_cols = []
    curr = start_date
    while len(date_cols) < 24:
        if curr.weekday() != 6: # Sunday is 6
            date_cols.append(curr.strftime('%b %d (%a)'))
        curr += datetime.timedelta(days=1)
        
    # Initialize the date matrix columns in df
    for date_col in date_cols:
        df[date_col] = 0
        
    # Solve daily spacing per merchandiser to balance store count per day
    for m in range(num_merch):
        merch_df = df[df['merch_id'] == m]
        if len(merch_df) == 0:
            continue
            
        # Daily workload count tracker (stores visited per day)
        merch_day_loads = np.zeros(24, dtype=int)
        
        # Sort outlets by frequency descending (place harder-to-space outlets first)
        sorted_indices = merch_df.sort_values(by='Actaul Frequency', ascending=False).index
        
        for idx in sorted_indices:
            freq = df.loc[idx, 'Actaul Frequency']
            
            if freq >= 24:
                # Visit all 24 days
                target_days = list(range(24))
            else:
                # Spacing solver: distribute evenly using floating spacing
                step_float = 24.0 / freq
                best_phase = []
                min_load = float('inf')
                
                # Test all possible starting phases [0, 23]
                for phase in range(24):
                    indices = sorted(list(set([int((phase + i * step_float) % 24) for i in range(freq)])))
                    if len(indices) < freq:
                        continue
                    
                    # Calculate cumulative daily store load for this candidate phase
                    load = sum(merch_day_loads[d_idx] for d_idx in indices)
                    if load < min_load:
                        min_load = load
                        best_phase = indices
                        
                if not best_phase:
                    best_phase = list(range(freq))
                    
                target_days = best_phase
                
            # Set binary visit indicators '1' and increment current workloads
            for d_idx in target_days:
                df.loc[idx, date_cols[d_idx]] = 1
                merch_day_loads[d_idx] += 1

    # ----------------------------------------------------
    # 4. OPENPYXL FORMATTED OUTPUT GENERATION
    # ----------------------------------------------------
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Tab 1: Executive Dashboard Summary
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Executive Dashboard Summary"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash.merge_cells("A1:E1")
    title_cell = ws_dash["A1"]
    title_cell.value = "PJP Balanced Workload Dashboard (24-Day Cycle)"
    title_cell.font = Font(name="Outfit", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 40
    
    # Headers
    headers_dash = [
        "Merchandiser Name", 
        "Total Outlets Assigned", 
        "Total Monthly Visits", 
        "Daily Store Load Range (Min-Max)", 
        "Average In-Store Hours/Day"
    ]
    for col_idx, h in enumerate(headers_dash, 1):
        cell = ws_dash.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = Font(name="Outfit", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_dash.row_dimensions[3].height = 28
    
    # Compute Dashboard Data
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    dash_row = 4
    for m in range(num_merch):
        merch_name = f"Merchandiser {m + 1:02d}"
        m_df = df[df['merch_id'] == m]
        
        total_outlets = len(m_df)
        total_visits = m_df['Actaul Frequency'].sum() if total_outlets > 0 else 0
        
        # Calculate daily workloads
        if total_outlets > 0:
            daily_visits = [m_df[date_col].sum() for date_col in date_cols]
            min_load = min(daily_visits)
            max_load = max(daily_visits)
            load_range = f"{min_load} to {max_load}"
        else:
            load_range = "0 to 0"
            
        # Average hours per day = Total Visits * 45 minutes / 60 / 24 working days
        # E.g. 10 visits/day * 45 mins = 450 mins = 7.5 hours
        avg_hours = (total_visits * 45.0) / 60.0 / 24.0
        
        # Style rows based on average hours
        if avg_hours > 8.0:
            fill_color = "FFD2D2" # Soft Red
            text_color = "9C0006"
        elif 6.5 <= avg_hours <= 8.0:
            fill_color = "FFF3CD" # Soft Yellow
            text_color = "9C6500"
        else:
            fill_color = "D4EDDA" # Soft Green
            text_color = "006100"
            
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        row_font = Font(name="Outfit", size=10, bold=False, color=text_color)
        
        cells = [
            ws_dash.cell(row=dash_row, column=1, value=merch_name),
            ws_dash.cell(row=dash_row, column=2, value=total_outlets),
            ws_dash.cell(row=dash_row, column=3, value=total_visits),
            ws_dash.cell(row=dash_row, column=4, value=load_range),
            ws_dash.cell(row=dash_row, column=5, value=round(avg_hours, 2))
        ]
        
        for idx, cell in enumerate(cells):
            cell.fill = row_fill
            cell.font = row_font
            cell.border = thin_border
            if idx == 0:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        ws_dash.row_dimensions[dash_row].height = 20
        dash_row += 1
        
    # Add an explanatory note
    ws_dash.merge_cells(start_row=dash_row+1, start_column=1, end_row=dash_row+2, end_column=5)
    note_cell = ws_dash.cell(row=dash_row+1, column=1)
    note_cell.value = "* Note: Average In-Store Hours/Day is calculated using a standard KSBCL field service metric of 45 minutes per outlet visit (Total Monthly Visits * 45 / 60 / 24). Rows are color-coded based on workload capacity: Red (>8h) indicates overload, Yellow (6.5h-8h) is target capacity, Green (<6.5h) represents available buffer capacity."
    note_cell.font = Font(name="Outfit", size=9, italic=True, color="555555")
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-adjust column widths for Dashboard
    for col in ws_dash.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # ----------------------------------------------------
    # Tab 2: Master PJP Matrix Grid
    # ----------------------------------------------------
    ws_grid = wb.create_sheet(title="Master PJP Matrix Grid")
    ws_grid.views.sheetView[0].showGridLines = True
    
    # Header list
    fixed_headers = ['Outlet code', 'Outlet name', 'Beat', 'LAT', 'LONG', 'Actaul Frequency', 'Assigned Merchandiser']
    all_headers = fixed_headers + date_cols
    
    # Set headers in row 1
    for col_idx, h in enumerate(all_headers, 1):
        cell = ws_grid.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(name="Outfit", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws_grid.row_dimensions[1].height = 32
    
    # Style definitions for data rows
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    bold_green_font = Font(name="Outfit", size=10, bold=True, color="155724")
    regular_font = Font(name="Outfit", size=10, color="333333")
    
    grid_row = 2
    for _, row in df.iterrows():
        is_even = (grid_row % 2 == 0)
        default_row_fill = zebra_fill if is_even else white_fill
        
        # Write fixed data columns
        fixed_values = [
            row['Outlet code'],
            row['Outlet name'],
            row['Beat'],
            row['LAT'],
            row['LONG'],
            row['Actaul Frequency'],
            row['Merchandiser Name']
        ]
        
        for col_idx, val in enumerate(fixed_values, 1):
            cell = ws_grid.cell(row=grid_row, column=col_idx, value=val)
            cell.fill = default_row_fill
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 6]: # Code, Frequency
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [4, 5]: # LAT, LONG
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        # Write 24 binary date columns
        for date_offset, date_col in enumerate(date_cols, len(fixed_values) + 1):
            val = int(row[date_col])
            cell = ws_grid.cell(row=grid_row, column=date_offset, value=val)
            cell.border = thin_border
            if val == 1:
                cell.fill = green_fill
                cell.font = bold_green_font
            else:
                cell.fill = default_row_fill
                cell.font = regular_font
                
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws_grid.row_dimensions[grid_row].height = 20
        grid_row += 1
        
    # Auto-adjust column widths for Grid
    for col in ws_grid.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        # Give date columns a standard small width
        if col[0].column > len(fixed_headers):
            ws_grid.column_dimensions[col_letter].width = 12
        else:
            ws_grid.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    if format_type == "excel":
        # Save to binary stream
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        return excel_stream.getvalue()
        
    elif format_type == "json":
        # Build JSON response structure for direct import into the portal
        outlets_list = []
        for _, row in df.iterrows():
            outlets_list.append({
                "Outlet code": str(int(row['Outlet code'])),
                "Outlet name": str(row['Outlet name']),
                "Beat": str(row['Beat']),
                "BM Mapped": "Business Manager",
                "Execution Champion Name": str(row['Merchandiser Name']),
                "Depot": "Default Depot",
                "Latitude": float(row['LAT']),
                "Longitude": float(row['LONG'])
            })
            
        bms = [{"name": "Business Manager", "role": "bm"}]
        ecs = []
        for m in range(num_merch):
            ecs.append({
                "name": f"Merchandiser {m + 1:02d}",
                "role": "ec",
                "bm": "Business Manager"
            })
        team_data = {"bms": bms, "ecs": ecs}
        
        DAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        pjp_data = {}
        
        for m in range(num_merch):
            merch_name = f"Merchandiser {m + 1:02d}"
            m_df = df[df['merch_id'] == m]
            if len(m_df) == 0:
                continue
                
            for d_idx, date_col in enumerate(date_cols):
                day_df = m_df[m_df[date_col] == 1]
                if len(day_df) == 0:
                    continue
                    
                week = (d_idx // 6) + 1
                day_name = DAYS[d_idx % 6]
                key = f"{week}__{merch_name.replace(' ', '_')}__{day_name}"
                
                # Sequence the outlets for this day using a nearest-neighbor TSP heuristic
                codes_to_sequence = day_df['Outlet code'].astype(int).tolist()
                sequenced_codes = []
                
                if len(codes_to_sequence) > 0:
                    # Extract active coordinates
                    lats = day_df['LAT'].values
                    lngs = day_df['LONG'].values
                    avg_lat = np.mean(lats)
                    avg_lng = np.mean(lngs)
                    
                    # Find closest to centroid average as start point
                    dists_to_avg = haversine_dist(lats, lngs, avg_lat, avg_lng)
                    start_idx = np.argmin(dists_to_avg)
                    
                    curr_idx = start_idx
                    visited = [curr_idx]
                    remaining = [i for i in range(len(day_df)) if i != curr_idx]
                    
                    while len(remaining) > 0:
                        last_lat = lats[curr_idx]
                        last_lng = lngs[curr_idx]
                        
                        best_rem_idx = 0
                        min_d = float('inf')
                        for r_idx in remaining:
                            d = haversine_dist(last_lat, last_lng, lats[r_idx], lngs[r_idx])
                            if d < min_d:
                                min_d = d
                                best_rem_idx = r_idx
                                
                        curr_idx = best_rem_idx
                        visited.append(curr_idx)
                        remaining.remove(curr_idx)
                        
                    sequenced_codes = [str(int(day_df.iloc[v_idx]['Outlet code'])) for v_idx in visited]
                    
                pjp_data[key] = sequenced_codes
                
        return {
            "outlets": outlets_list,
            "team": team_data,
            "pjp": pjp_data
        }
